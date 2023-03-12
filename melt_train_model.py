import xgboost as xgb
import lightgbm as gbm
import pandas as pd
import numpy as np
import openpyxl
import argparse
import pickle
import os
import melt_constants as const

from sklearn import metrics

# load workpiece and material property data from excel file
def load_data(workpiece_filepath, property_filepath, keys, header=[]):
    x_train = {key: [] for key in keys}
    x_test = {key: [] for key in keys}
    y_train = {key: [] for key in keys}
    y_test = {key: [] for key in keys}
    remove_header = const.layer_header + header

    for wpf, ppf in zip(workpiece_filepath, property_filepath):
        print('read workpiece file: ', wpf)
        print('read material property file: ', ppf)
        wp_sheets = pd.read_excel(wpf, sheet_name=None)

        for key in keys:
            # the dictionary key must same in property excel
            pp_data = pd.read_excel(ppf, sheet_name=key)
            pp_data = np.array(pp_data.drop(const.trail_header, axis=1))
            # switch 2D array to 1D array
            pp_data = np.reshape(pp_data, -1)

            for index, sheet in zip(range(len(pp_data)), wp_sheets):
                wp_data = pd.read_excel(wpf, sheet_name=sheet)
                if pp_data[index] == const.error_data:
                    continue

                if sheet[-1] == '5' or sheet[-1] == '6':
                    x_test[key].append(wp_data.drop(remove_header, axis=1))
                    y_test[key].append(pp_data[index])
                else:
                    x_train[key].append(wp_data.drop(remove_header, axis=1))
                    y_train[key].append(pp_data[index])

    return x_train, x_test, y_train, y_test

def store_data(sheet, x_train, x_test, y_test, model, predict):
    for col, header in zip(range(1, len(const.output_header) + 1),
                           const.output_header):
        sheet.cell(1, col).value = header
    sheet.cell(2, 4).value = x_train.shape[0]
    sheet.cell(2, 5).value = x_test.shape[0]
    # compute R2 score, MSE and MAE
    sheet.cell(2, 6).value = model.score(x_test, y_test)
    sheet.cell(2, 7).value = \
        metrics.mean_squared_error(y_test, predict)
    sheet.cell(2, 8).value = \
        metrics.mean_absolute_error(y_test, predict)

    row = 2
    for pre, true in zip(predict, y_test):
        sheet.cell(row, 1).value = pre
        sheet.cell(row, 2).value = true
        sheet.cell(row, 3).value = (abs(pre - true) / true) * 100
        row += 1

# data preprocessing to train and test data
def data_preprocessing(x_train, x_test, y_train, y_test, x_shape, y_repeat):
    for key in x_train.keys():
        # convert 3-D matrix to 2-D matrix
        x_train[key] = np.reshape(np.array(x_train[key]), x_shape)
        x_test[key] = np.reshape(np.array(x_test[key]), x_shape)
        y_train[key] = np.repeat(y_train[key], y_repeat)
        y_test[key] = np.repeat(y_test[key], y_repeat)
    return x_train, x_test, y_train, y_test

# display the correlation coefficient
def print_corrcoef(x_train, x_test, y_train, y_test):
    for key in x_train.keys():
        x = np.vstack((x_train[key], x_test[key]))
        y = np.concatenate((y_train[key], y_test[key]))
        x_mean = np.mean(x, axis=0)
        y_mean = np.mean(y)

        numerator = np.zeros(shape=x_mean.shape[0])
        denominator1 = np.zeros(shape=x_mean.shape[0])
        denominator2 = np.zeros(shape=x_mean.shape[0])

        # compute correlation coefficient
        for xi, yi in zip(x, y):
            x_tmp = xi - x_mean
            y_tmp = yi - y_mean
            numerator += x_tmp * y_tmp
            denominator1 += x_tmp ** 2
            denominator2 += y_tmp ** 2

        r = numerator / (np.sqrt(denominator1) * np.sqrt(denominator2))
        print('----------', key, '----------')
        for feature, i in zip(const.feature_header, r):
            print(feature, ': ', i)

def train_xgboost_model(x_train, x_test, y_train, y_test,
                        keys, output, xlsx):
    wb = openpyxl.Workbook()

    for key in keys:
        model = xgb.XGBRegressor(
            n_estimators=10000, learning_rate=0.3, max_depth=5)
        # train XGBoost model
        model.fit(x_train[key], y_train[key])
        # save XGBoost model
        pickle.dump(model, open(output + key + '.pickle.dat', 'wb'))
        # predict x_test via XGBoost model
        predict = model.predict(x_test[key])

        # create new sheet
        sheet = wb.create_sheet(key)
        # store data into excel
        store_data(sheet, x_train[key], x_test[key],
                   y_test[key], model, predict)
        # save the excel
        wb.save(output + xlsx)

def train_lightgbm_model(x_train, x_test, y_train, y_test,
                         keys, output, xlsx):
    wb = openpyxl.Workbook()

    for key in keys:
        model = gbm.LGBMRegressor(boosting_type='gbdt', num_leaves=10000,
                                  learning_rate=0.3, max_depth=6)
        # train lightGBM model
        model.fit(x_train[key], y_train[key])
        # save lightGBM model
        pickle.dump(model, open(output + key + '.pickle.dat', 'wb'))
        # predict x_test via lightGBM model
        predict = model.predict(x_test[key])

        # create new sheet
        sheet = wb.create_sheet(key)
        # store data into excel
        store_data(sheet, x_train[key], x_test[key],
                   y_test[key], model, predict)
        # save the excel
        wb.save(output + xlsx)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--dst',
                        default='./result/xgboost/',
                        help='destination path')
    args = parser.parse_args()

    if not os.path.isdir(args.dst):
        os.makedirs(args.dst)

    bone_filepath = ['./data/glcm-data/第一批狗骨頭.xlsx',
                     './data/glcm-data/第二批狗骨頭.xlsx']

    bone_property_filepath = ['./data/glcm-data/第一批狗骨頭材料特性.xlsx',
                              './data/glcm-data/第二批狗骨頭材料特性.xlsx']

    ring_filepath = ['./data/glcm-data/第一批圓環.xlsx',
                     './data/glcm-data/第二批圓環.xlsx']

    ring_property_filepath = ['./data/glcm-data/第一批圓環材料特性.xlsx',
                              './data/glcm-data/第二批圓環材料特性.xlsx']

    # load tensile data
    x_train, x_test, y_train, y_test = load_data(bone_filepath,
                                                 bone_property_filepath,
                                                 const.tensile_key)
    # data preprocessing
    x_train, x_test, y_train, y_test = \
        data_preprocessing(x_train, x_test, y_train, y_test,
                           x_shape=(-1, 70 * 13), y_repeat=70)
    # train tensile model
    train_xgboost_model(x_train, x_test, y_train, y_test,
                        keys=const.tensile_key, output=args.dst,
                        xlsx='tensile.xlsx')

    # load permeability data
    x_train, x_test, y_train, y_test = load_data(ring_filepath,
                                                 ring_property_filepath,
                                                 const.pmb_key)
    # data preprocessing
    x_train, x_test, y_train, y_test = \
        data_preprocessing(x_train, x_test, y_train, y_test,
                           x_shape=(-1, 13), y_repeat=70)
    # train permeability model
    train_xgboost_model(x_train, x_test, y_train, y_test,
                        keys=const.pmb_key, output=args.dst,
                        xlsx='permeability.xlsx')

    # load iron loss data
    x_train, x_test, y_train, y_test = load_data(ring_filepath,
                                                 ring_property_filepath,
                                                 const.iron_key)
    # data preprocessing
    x_train, x_test, y_train, y_test = \
        data_preprocessing(x_train, x_test, y_train, y_test,
                           x_shape=(-1, 13), y_repeat=70)
    # train iron loss model
    train_xgboost_model(x_train, x_test, y_train, y_test,
                        keys=const.iron_key, output=args.dst,
                        xlsx='iron.xlsx')
